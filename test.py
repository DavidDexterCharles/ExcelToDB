from openpyxl import load_workbook
import re
# import json
import simplejson as json
from flask import Flask, jsonify
from flask_sqlalchemy import SQLAlchemy


app = Flask(__name__)
app.config.from_pyfile('config.cfg')
db = SQLAlchemy(app)

#https://www.datacamp.com/community/tutorials/python-excel-tutorial#gs.RwPnpzg

# Load in the workbook
wb = load_workbook('test.xlsx')

# Get sheet names
# print wb.get_sheet_names()

sheet = wb.get_sheet_by_name('sheet1')
db.engine.execute('select * from infoho')
# Database aspect
# http://www.python-course.eu/sql_python.php
# https://stackoverflow.com/questions/23110383/how-to-dynamically-build-a-json-object-with-python

# print sheet['B2'].value

# print len(sheet['B4:J7'][0])
# print str (sheet['B4:J7'][0][0].value) +"  "+ str (sheet['B4:J7'][0][1].value)
# print str (sheet['B4:J7'][1][0].value) +"  "+ str (sheet['B4:J7'][1][1].value)
# print str (sheet['B4:J7'][2][0].value) +"  "+ str (sheet['B4:J7'][2][1].value)

# print "\n\n"


# print len(sheet['A5:J7'])
# print str (sheet['A5:J7'][0][0].value) +"  "+  str (sheet['A5:J7'][1][0].value)
# print str (sheet['A5:J7'][0][1].value) +"  "+  str (sheet['A5:J7'][1][1].value)
# print str (sheet['A5:J7'][0][2].value) +"  "+ str (sheet['A5:J7'][1][2].value)


class Tools(object):

   def getjsondata(self,path):
        with open(path) as data_file:    
            data = json.load(data_file)
        return data


class ExceltoDBAPI(Tools):

    def __init__(self):
        self.mapdb = {}

    def gettables(self, path):
        
        tables= self.getjsondata(path)["tables"]
        # print tables
        for x in range(0,len(tables)):
            t={}
            t["tablename"]=tables[x]["name"]
            t["tablerange"]=tables[x]["table"]
            t["tableorient"]=tables[x]["orient"]
            
            self.get_one_table(x,tables[x]["table"],tables[x]["orient"],t)
        # print len(self.mapdb[0]['headers'][0]['data'])
        # print self.mapdb[0]['tablename'].lower()

        print self.mapdb
        tablename = self.mapdb[0]['tablename'].lower() # the table name also note for headers, the index starts at 0
        headers = self.mapdb[0]['headers']
        header = self.mapdb[0]['headers'][1]['h'] #header1
        data = self.mapdb[0]['headers'][1]['data'] #data for header 1
        datacount = len(self.mapdb[0]['headers'][1]['data']) #amount of data for header 1
        self.insertinto(tablename,headers)
        # print  json.dumps(self.mapdb)

    def insertinto(self,tablename,headers):
        print tablename
        # print len(headers[0]["data"][0])
        # print headers[0]["data"][0]
        # print headers
        # print len(headers)
        headerslist=""
        rowlist={}
        hldata={} #header list data
        for i in range(0,len(headers)):
            header = headers[i]['h']
            headerslist += header+", "
            hldata[i]={}
            rowdata=""
            for j in range(1, len(headers[i]["data"])):
                hldata[i][j]=headers[i]['data'][j][j]
                # rowdata+= "'"+headers[i]['data'][j][j]+"', "
                # rowlist[temp]= rowdata.rstrip(', ')
                # temp
            # print header 


        # print hldata[0]
        # print hldata[1]
        # print hldata[5]
        # rowlist[x]={}
        ch={}
        for x in range (1,len(hldata[0])+1):
            ch[x]=""
        for y in range(0,len(hldata)):
            # print hldata[y]
            # ch={}
            for x in range (1,len(hldata[y])+1):
                # print hldata[y][x] + ","
                ch[x]+="'"+hldata[y][x] + "', "
            # print ch
            # print "\n"
        for x in range (1,len(hldata[0])+1):
            ch[x]=ch[x].rstrip(', ')  
        # print ch
        # print headerslist
        # for x in range (1,len(ch)+1):
        #     db.engine.execute("INSERT INTO "+tablename+" ("+headerslist.rstrip(', ')  +") VALUES ("+ch[x]+")")
       
        # DELETE FROM `infoho` WHERE 1


    def get_one_table(self,tindex,table,orient,t):
        
        tablerange = table.replace("-", ":")
        if orient=="vertical" or orient=="v":
            headercount = len(sheet[tablerange])
        else: # Orientation is horizontal [columns left to right]
            headercount = len(sheet[tablerange][0])
            vcount = len(sheet[tablerange])
            t["headercount"]=str(headercount)
            t["vcount"]=str(vcount)
            t["headers"]={}
            headers={}
            h={}
            for i in range(0,headercount):
                attrval = sheet[tablerange][0][i].value
                header={}
                header["h"]=attrval
                header["data"]={}
                data={}
              

                # print attrval
                for j in range(0,vcount):
                    
                    # print sheet[tablerange][j][i].value
                    d={}
                    d[j]=str(sheet[tablerange][j][i].value)
                    data[j]= d
                header["data"]=data
                headers[i]=header
            t["headers"]=headers   
            self.mapdb[tindex] = t
            
        # print tablerange
        # print headercount

        







api = ExceltoDBAPI()
api.gettables("tables.json")
'''
# Retrieve cell value 
sheet.cell(row=1, column=2).value

# Print out values in column 2 
for i in range(1, 30):
    print i
    for j in range(1, 24):
        val = sheet.cell(row=i, column=j).value
        # if re.search("None" , str(val)) and r:
        #     c=j
        #     r=i
        
        if not re.search("None" , str(val)):
            print val

'''